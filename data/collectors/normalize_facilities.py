"""편의시설 데이터 폴더 정리 + 컬럼 표준화 + 필수 속성 검증 → REPORT.md 생성"""

import csv
import re
import shutil
from pathlib import Path
from datetime import datetime

import openpyxl

ROOT = Path(__file__).resolve().parents[1] / "편의시설"

# 표준 통일 컬럼 (DB living_infra_gumi 스키마 매핑 친화)
STD_COLS = [
    "category", "sub_category", "name", "lat", "lon",
    "address", "phone", "operator", "source_type", "source_id", "data_source",
]

# DB CHECK 제약상 허용되는 enum
ALLOWED_CATEGORIES = {"medical", "education", "convenience", "transport", "safety", "leisure"}
ALLOWED_SOURCES = {"kakao", "gumi_opendata", "data_go_kr", "manual"}

# 카카오 카테고리 분류명을 보고 sub_category 세분화 (어린이집/유치원 등)
def refine_kakao_kindergarten(category_name: str) -> str:
    if "어린이집" in category_name:
        return "daycare"
    if "유치원" in category_name:
        return "kindergarten"
    return "kindergarten_or_daycare"


def read_csv(path: Path) -> list[dict]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path: Path, rows: list[dict]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=STD_COLS)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in STD_COLS})


def standardize_osm(rows: list[dict], category: str, sub_category: str) -> list[dict]:
    out = []
    for r in rows:
        out.append({
            "category": category,
            "sub_category": sub_category,
            "name": r.get("name", "").strip(),
            "lat": r.get("lat", ""),
            "lon": r.get("lon", ""),
            "address": r.get("address", ""),
            "phone": r.get("phone", ""),
            "operator": r.get("operator", ""),
            "source_type": "osm",
            "source_id": f'{r.get("osm_type","")}/{r.get("osm_id","")}'.strip("/"),
            "data_source": "manual",  # DB enum에 osm 없음 → manual
        })
    return out


def standardize_kakao(rows: list[dict], category: str, sub_category: str, sub_refiner=None) -> list[dict]:
    out = []
    for r in rows:
        final_sub = sub_category
        if sub_refiner:
            final_sub = sub_refiner(r.get("category_name", "")) or sub_category
        out.append({
            "category": category,
            "sub_category": final_sub,
            "name": r.get("name", "").strip(),
            "lat": r.get("lat", ""),
            "lon": r.get("lon", ""),
            "address": r.get("address", ""),
            "phone": r.get("phone", ""),
            "operator": r.get("operator", ""),
            "source_type": "kakao",
            "source_id": str(r.get("kakao_id", "")),
            "data_source": "kakao",
        })
    return out


def convert_bus_stop_xlsx() -> list[dict]:
    src = ROOT / "구미시_버스정류장_전체데이터O_20260513_101636.xlsx"
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb.active
    out = []
    seen = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            # headers: sttn_idntfr, sttn_std_idntfr, sttn_ty, centr_cartrk_se,
            # korean_nm, shrten_nm, la, lo, ...
            continue
        if not row or row[0] is None:
            continue
        sttn_id = row[0]
        name = (row[4] or "").strip()
        # 컬럼명 함정: la=경도, lo=위도
        lon = row[6]
        lat = row[7]
        if lat is None or lon is None or not name:
            continue
        key = (str(sttn_id), name)
        if key in seen:
            continue
        seen.add(key)
        out.append({
            "category": "transport",
            "sub_category": "bus_stop",
            "name": name,
            "lat": lat,
            "lon": lon,
            "address": "",
            "phone": "",
            "operator": "",
            "source_type": "gumi_opendata_xlsx",
            "source_id": str(sttn_id),
            "data_source": "gumi_opendata",
        })
    return out


def convert_cctv_xlsx() -> list[dict]:
    src = ROOT / "구미시_생활방범cctv_전체데이터O_20260513_102038.xlsx"
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb.active
    out = []
    seen = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            # headers: sn, emd_nm, old_adres, nw_adres, detail_lc, camera_qy, manage_no
            continue
        if not row or row[0] is None:
            continue
        sn = row[0]
        emd = (row[1] or "").strip()
        old_addr = (row[2] or "").strip()
        new_addr = (row[3] or "").strip()
        detail_loc = (row[4] or "").strip()
        manage_no = (row[6] or "").strip()
        address = new_addr or old_addr
        name = manage_no if manage_no else f"CCTV-{sn}"
        key = (str(sn), address)
        if key in seen:
            continue
        seen.add(key)
        out.append({
            "category": "safety",
            "sub_category": "cctv",
            "name": name,
            "lat": "",  # 원본에 좌표 없음 - 지오코딩 필요
            "lon": "",
            "address": f"{emd} {address} {detail_loc}".strip(),
            "phone": "",
            "operator": "",
            "source_type": "gumi_opendata_xlsx",
            "source_id": str(sn),
            "data_source": "gumi_opendata",
        })
    return out


def is_number(v) -> bool:
    if v is None or v == "":
        return False
    try:
        float(v)
        return True
    except (ValueError, TypeError):
        return False


def validate(rows: list[dict]) -> dict:
    n = len(rows)
    miss_lat = sum(1 for r in rows if not is_number(r.get("lat")))
    miss_lon = sum(1 for r in rows if not is_number(r.get("lon")))
    miss_name = sum(1 for r in rows if not (r.get("name") or "").strip())
    miss_addr = sum(1 for r in rows if not (r.get("address") or "").strip())
    bad_cat = sum(1 for r in rows if r.get("category") not in ALLOWED_CATEGORIES)
    bad_src = sum(1 for r in rows if r.get("data_source") not in ALLOWED_SOURCES)
    # 한국 범위 위경도 sanity check (위도 33~39, 경도 124~132)
    bad_coord = 0
    for r in rows:
        if is_number(r.get("lat")) and is_number(r.get("lon")):
            la = float(r["lat"]); lo = float(r["lon"])
            if not (33 <= la <= 39 and 124 <= lo <= 132):
                bad_coord += 1
    return {
        "total": n,
        "missing_lat": miss_lat,
        "missing_lon": miss_lon,
        "missing_name": miss_name,
        "missing_address": miss_addr,
        "invalid_category": bad_cat,
        "invalid_data_source": bad_src,
        "out_of_korea_bbox": bad_coord,
    }


# 소스 → 목적지 매핑
LAYOUT = [
    # (source_path_rel, dst_rel, category, sub_category, source_kind, sub_refiner)
    ("osm/medical_hospital.csv",                 "medical/hospital.csv",          "medical",     "hospital",          "osm",   None),
    ("kakao/medical_pharmacy.csv",               "medical/pharmacy.csv",          "medical",     "pharmacy",          "kakao", None),
    ("osm/education_school_all.csv",             "education/school.csv",          "education",   "school",            "osm",   None),
    ("osm/education_school_elementary.csv",      "education/school_elementary_guess.csv", "education", "elementary_school", "osm", None),
    ("kakao/education_kindergarten.csv",         "education/kindergarten.csv",    "education",   "kindergarten",      "kakao", refine_kakao_kindergarten),
    ("kakao/education_academy.csv",              "education/academy.csv",         "education",   "academy",           "kakao", None),
    ("osm/convenience_store.csv",                "convenience/convenience_store.csv", "convenience", "convenience_store", "osm", None),
    ("osm/convenience_mart.csv",                 "convenience/mart.csv",          "convenience", "mart",              "osm",   None),
    ("osm/safety_police.csv",                    "safety/police.csv",             "safety",      "police_station",    "osm",   None),
    ("osm/leisure_park.csv",                     "leisure/park.csv",              "leisure",     "park",              "osm",   None),
    ("osm/leisure_library.csv",                  "leisure/library.csv",           "leisure",     "library",           "osm",   None),
    ("osm/leisure_cultural.csv",                 "leisure/cultural.csv",          "leisure",     "cultural_center",   "osm",   None),
]


def merge_kakao_fire() -> list[dict]:
    """카카오 119안전센터 + 소방서 두 CSV를 합쳐서 fire_station으로 단일화."""
    out = []
    seen_ids = set()
    for fname in ("kakao/safety_fire_safety_center.csv", "kakao/safety_fire_station.csv"):
        p = ROOT / fname
        if not p.exists():
            continue
        for r in read_csv(p):
            kid = str(r.get("kakao_id", ""))
            if kid in seen_ids:
                continue
            seen_ids.add(kid)
            out.append(r)
    return standardize_kakao(out, "safety", "fire_station")


def main():
    print(f"[setup] working under {ROOT}")
    raw_dir = ROOT / "raw"
    archive_dir = ROOT / "archive"
    raw_dir.mkdir(exist_ok=True)
    archive_dir.mkdir(exist_ok=True)

    # 1) 원본 xlsx/docx → raw/로 이동 (이미 들어가 있으면 패스)
    for fname in [
        "구미시_버스정류장_전체데이터O_20260513_101636.xlsx",
        "구미시_생활방범cctv_전체데이터O_20260513_102038.xlsx",
        "Bus_Stop_Variable_Descriptions.docx",
        "생활 방범CCTV_Definitions.docx",
    ]:
        src = ROOT / fname
        if src.exists():
            dst = raw_dir / fname
            print(f"[raw] move {fname} -> raw/")
            shutil.move(str(src), str(dst))

    # 2) xlsx 변환 (raw 위치 사용)
    bus_rows = convert_bus_stop_xlsx_from_raw(raw_dir)
    cctv_rows = convert_cctv_xlsx_from_raw(raw_dir)
    write_csv(ROOT / "transport/bus_stop.csv", bus_rows)
    write_csv(ROOT / "safety/cctv.csv", cctv_rows)
    print(f"[xlsx] transport/bus_stop.csv: {len(bus_rows)} rows")
    print(f"[xlsx] safety/cctv.csv: {len(cctv_rows)} rows (좌표 없음, 지오코딩 필요)")

    # 3) osm/kakao CSV → 카테고리 폴더 (표준화)
    report = []
    for src_rel, dst_rel, cat, sub, kind, refiner in LAYOUT:
        src = ROOT / src_rel
        if not src.exists():
            print(f"[skip] {src_rel} not found")
            continue
        rows_raw = read_csv(src)
        if kind == "osm":
            std = standardize_osm(rows_raw, cat, sub)
        else:
            std = standardize_kakao(rows_raw, cat, sub, refiner)
        write_csv(ROOT / dst_rel, std)
        print(f"[std] {src_rel} -> {dst_rel}: {len(std)} rows")

    # 소방서 (카카오 두 파일 통합)
    fire_rows = merge_kakao_fire()
    write_csv(ROOT / "safety/fire_station.csv", fire_rows)
    print(f"[std] kakao fire merged -> safety/fire_station.csv: {len(fire_rows)} rows")

    # 4) 미사용 원본 csv → archive/
    for src_rel in [
        "osm/medical_pharmacy.csv",      # 카카오 146건 사용, OSM 5건 archive
        "osm/education_kindergarten.csv",  # 카카오 295건 사용, OSM 15건 archive
        "osm/education_academy.csv",     # 카카오 529건 사용, OSM 0건 archive
        "osm/safety_fire.csv",           # 카카오 14건 사용, OSM 1건 archive
        "kakao/safety_fire_station.csv", # 통합 사용
        "kakao/safety_fire_safety_center.csv",  # 통합 사용
    ]:
        src = ROOT / src_rel
        if not src.exists():
            continue
        dst = archive_dir / src_rel.replace("/", "_")
        print(f"[archive] {src_rel} -> archive/{dst.name}")
        shutil.move(str(src), str(dst))

    # 빈 osm/, kakao/ 폴더 제거
    for empty in ("osm", "kakao"):
        d = ROOT / empty
        if d.exists() and not any(d.iterdir()):
            d.rmdir()
            print(f"[clean] removed empty {empty}/")
        elif d.exists():
            # 남은 파일이 있으면 archive로 통합
            for f in list(d.iterdir()):
                target = archive_dir / f"{empty}_{f.name}"
                print(f"[archive] {empty}/{f.name} -> archive/{target.name}")
                shutil.move(str(f), str(target))
            d.rmdir()
            print(f"[clean] removed {empty}/")

    # 5) 모든 최종 CSV 검증
    print("\n[validate] scanning final CSVs...")
    final_files = sorted([p for p in ROOT.rglob("*.csv") if "archive" not in p.parts])
    summary = []
    for p in final_files:
        rows = read_csv(p)
        v = validate(rows)
        rel = p.relative_to(ROOT)
        summary.append((str(rel).replace("\\", "/"), v))

    # 6) REPORT.md 생성
    md = ["# 편의시설 데이터 검증 리포트", "",
          f"생성 시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", "",
          "## 표준 컬럼", f"`{', '.join(STD_COLS)}`", "",
          "## DB 적재 필수 속성",
          "- `category` (CHECK enum: medical/education/convenience/transport/safety/leisure)",
          "- `sub_category` (자유 문자열)",
          "- `name` (NOT NULL)",
          "- `lat`, `lon` (`point_geom`으로 변환, NOT NULL)",
          "- `data_source` (CHECK enum: kakao/gumi_opendata/data_go_kr/manual)",
          "",
          "## 파일별 검증 결과", "",
          "| 파일 | 총 행 | lat 누락 | lon 누락 | name 누락 | address 누락 | 좌표범위이상 |",
          "|---|---:|---:|---:|---:|---:|---:|"]
    grand_total = 0
    missing_coords_files = []
    for rel, v in summary:
        md.append(f"| `{rel}` | {v['total']} | {v['missing_lat']} | {v['missing_lon']} | {v['missing_name']} | {v['missing_address']} | {v['out_of_korea_bbox']} |")
        grand_total += v["total"]
        if v["missing_lat"] > 0 or v["missing_lon"] > 0:
            missing_coords_files.append((rel, v))
    md.append(f"\n**합계: {grand_total}건**\n")

    # 필수 속성 누락 데이터 + 보충 방안
    md.append("## 🔴 필수 속성 누락 데이터\n")
    if not missing_coords_files:
        md.append("좌표(lat/lon) 누락 행 없음.")
    else:
        md.append("### 좌표 누락\n")
        for rel, v in missing_coords_files:
            md.append(f"- **`{rel}`**: lat 누락 {v['missing_lat']}건 / lon 누락 {v['missing_lon']}건 (총 {v['total']}건 중)")
        md.append("")
        md.append("### 보충 방안")
        md.append("")
        md.append("**1. `safety/cctv.csv` (940건 좌표 없음)**")
        md.append("- 원본(`raw/구미시_생활방범cctv_*.xlsx`)에 도로명/지번 주소만 있음")
        md.append("- 보충 출처:")
        md.append("  - **VWorld 주소 검색 API** ([`GeocoderClient`](../../backend/src/main/java/com/arproperty/external/vworld/GeocoderClient.java)) — 무료, 백엔드에 클라이언트 스켈레톤 존재")
        md.append("  - **카카오 주소→좌표 변환 API** `/v2/local/search/address.json` — 일일 30만회 무료, 이미 키 등록됨")
        md.append("  - **공공데이터포털 행정안전부 주소→좌표 API** [#15057559](https://www.data.go.kr/data/15057559/openapi.do)")
        md.append("- 작업량: 940건 × 단건 호출 = 약 1~2분 (카카오 RPS 10 기준)")
        md.append("")
    md.append("## 🟡 데이터 품질 경고\n")
    md.append("- `medical/hospital.csv` (OSM, 217건): 비의료 시설 일부 섞임(예: '영신문방구'). DB 적재 전 `name` LIKE 필터 수동 검수 권장")
    md.append("- `education/school_elementary_guess.csv` (OSM, 46건): OSM의 `school:type` 태그 누락으로 초/중/고 구분 정확도 낮음. 학교명 후처리로 분류 권장")
    md.append("- `convenience/convenience_store.csv` (OSM, 70건): 실제 구미시 편의점 200+개 추정. 카카오 `CS2` 카테고리로 보충 가능")
    md.append("")
    md.append("## 📂 최종 폴더 구조")
    md.append("```")
    md.append("data/편의시설/")
    md.append("├── raw/                        # 원본 xlsx/docx 보존")
    md.append("├── archive/                    # 사용 안 하는 보조 데이터")
    md.append("├── medical/")
    md.append("│   ├── hospital.csv            (OSM)")
    md.append("│   └── pharmacy.csv            (Kakao)")
    md.append("├── education/")
    md.append("│   ├── school.csv              (OSM)")
    md.append("│   ├── school_elementary_guess.csv  (OSM)")
    md.append("│   ├── kindergarten.csv        (Kakao)")
    md.append("│   └── academy.csv             (Kakao)")
    md.append("├── convenience/")
    md.append("│   ├── convenience_store.csv   (OSM)")
    md.append("│   └── mart.csv                (OSM)")
    md.append("├── transport/")
    md.append("│   └── bus_stop.csv            (xlsx → CSV)")
    md.append("├── safety/")
    md.append("│   ├── police.csv              (OSM)")
    md.append("│   ├── fire_station.csv        (Kakao)")
    md.append("│   └── cctv.csv                (xlsx, 좌표 없음)")
    md.append("├── leisure/")
    md.append("│   ├── park.csv                (OSM)")
    md.append("│   ├── library.csv             (OSM)")
    md.append("│   └── cultural.csv            (OSM)")
    md.append("└── REPORT.md")
    md.append("```")

    (ROOT / "REPORT.md").write_text("\n".join(md), encoding="utf-8")
    print(f"\n[report] {ROOT/'REPORT.md'} written ({len(md)} lines)")


def convert_bus_stop_xlsx_from_raw(raw_dir: Path) -> list[dict]:
    src = raw_dir / "구미시_버스정류장_전체데이터O_20260513_101636.xlsx"
    if not src.exists():
        return []
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb.active
    out = []
    seen = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or row[0] is None:
            continue
        sttn_id = row[0]
        name = (row[4] or "").strip()
        lon = row[6]  # la 컬럼 = 경도
        lat = row[7]  # lo 컬럼 = 위도
        if lat is None or lon is None or not name:
            continue
        key = (str(sttn_id), name)
        if key in seen:
            continue
        seen.add(key)
        out.append({
            "category": "transport",
            "sub_category": "bus_stop",
            "name": name,
            "lat": lat,
            "lon": lon,
            "address": "",
            "phone": "",
            "operator": "",
            "source_type": "gumi_opendata_xlsx",
            "source_id": str(sttn_id),
            "data_source": "gumi_opendata",
        })
    return out


def convert_cctv_xlsx_from_raw(raw_dir: Path) -> list[dict]:
    src = raw_dir / "구미시_생활방범cctv_전체데이터O_20260513_102038.xlsx"
    if not src.exists():
        return []
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb.active
    out = []
    seen = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or row[0] is None:
            continue
        sn = row[0]
        emd = (row[1] or "").strip()
        old_addr = (row[2] or "").strip()
        new_addr = (row[3] or "").strip()
        detail_loc = (row[4] or "").strip()
        manage_no = (row[6] or "").strip()
        address = new_addr or old_addr
        name = manage_no if manage_no else f"CCTV-{sn}"
        key = (str(sn), address)
        if key in seen:
            continue
        seen.add(key)
        out.append({
            "category": "safety",
            "sub_category": "cctv",
            "name": name,
            "lat": "",
            "lon": "",
            "address": f"{emd} {address} {detail_loc}".strip(),
            "phone": "",
            "operator": "",
            "source_type": "gumi_opendata_xlsx",
            "source_id": str(sn),
            "data_source": "gumi_opendata",
        })
    return out


if __name__ == "__main__":
    main()
